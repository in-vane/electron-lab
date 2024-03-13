import os
import re
import csv
import shutil
import base64
from io import BytesIO

import cv2
import fitz
import pandas as pd
from PIL import Image
from tabula import read_pdf
from collections import defaultdict
from ppocronnx.predict_system import TextSystem
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font



CURRENT_PATH = os.path.dirname(os.path.abspath(__file__))
PDF_PATH = os.path.join(CURRENT_PATH, "temp.pdf")
IMAGE_PATH = os.path.join(CURRENT_PATH, "image")
CSV_PATH = os.path.join(CURRENT_PATH, "selected_table.csv")


def find_target_table(doc):
    total_pages = len(doc)  # 获取PDF的总页数
    
    # 遍历每一页
    for page_num in range(1, total_pages + 1):
        # 使用tabula读取当前页的表格
        df_list = read_pdf(PDF_PATH, pages=page_num, multiple_tables=True)

        for df in df_list:
            # 检查表头和内容是否符合预期
            if all(x in df.columns for x in ['A', 'B', 'C']) and 'x' in ''.join(df.iloc[:, 1].astype(str)):
                df.to_csv(CSV_PATH, index=False) # 找到符合条件的表格，保存为CSV文件
                return page_num


# 处理每个单元格数据
def clean_cell(cell):
    if isinstance(cell, str):
        # 移除"x"并保留数字
        cell = re.sub(r'x(\d+)', r'\1', cell)
        # 保留只包含一个大写字母或数字的单元格内容
        if re.match(r'^[A-Z]$', cell) or re.match(r'^\d+$', cell):
            return cell
    return None


# 获取大写英问字符和对应数字
def manage_csv():
    # 尝试读取CSV文件，假设它位于可以访问的路径
    try:
        df = pd.read_csv(CSV_PATH, header=None)
        # 应用清理函数到每个单元格
        df = df.applymap(clean_cell)
        # 删除全为空的列
        df.dropna(axis=1, how='all', inplace=True)
        # 覆盖原来的CSV文件
        df.to_csv(CSV_PATH, index=False, header=False)
    except Exception:
        pass


# csv转字典
def read_csv_to_dict():
    result_dict = {}

    with open(CSV_PATH, 'r', encoding='utf-8') as csv_file:
        # 使用 csv.reader 读取 CSV 文件
        csv_reader = csv.reader(csv_file)

        # 读取第一行为字符，第二行为数字
        characters = next(csv_reader)
        numbers = next(csv_reader)

        # 清理 numbers 列表，确保只包含数字
        cleaned_numbers = []
        for item in numbers:
            # 提取字符串中的数字部分
            match = re.search(r'\d+', item)
            if match:
                cleaned_numbers.append(int(match.group(0)))
            else:
                cleaned_numbers.append(0)  # 如果没有找到数字，使用0作为默认值


        # 将字符和数字对应存储到字典中
        result_dict = dict(zip(characters, cleaned_numbers))

    return result_dict


# 获取总的螺丝表
def get_total_screw(doc):
    page_num = find_target_table(doc)
    manage_csv()
    result_dict = read_csv_to_dict()

    return result_dict, page_num


# 提取步骤图里的螺丝表
def extract_images_below_steps(doc):
    # 步骤编号的正则表达式，匹配大于0的数字后跟一个点
    step_pattern = re.compile(r'(?<!\.\d)\b(?<!\d\.)[1-9]\d*\.(?!\d)')

    extracted_images = []

    for page_num in range(len(doc)):
        page = doc[page_num]
        # 在Python层面上使用正则表达式找到所有步骤编号
        for match in re.finditer(step_pattern, page.get_text("text")):
            # 转换匹配的文本为坐标
            match_rects = page.search_for(match.group(0))
            for rect in match_rects:
                # 定义步骤编号下方的矩形区域，此处可能需要调整
                clip_rect = fitz.Rect(rect.x0 - 3, rect.y1 - 20, rect.x1 + 800, rect.y1 + 315)
                # 提取该区域的图像
                pix = page.get_pixmap(clip=clip_rect)
                # 定义图像的保存路径
                image_filename = f"step_{page_num + 1}_{int(rect.x0)}_{int(rect.y1)}.png"
                image_filepath = os.path.join(IMAGE_PATH, image_filename)
                # 保存图像
                pix.save(image_filepath)
                extracted_images.append(image_filepath)

    return extracted_images


def recognize_text_in_images(image_paths):
    # Initialize TextSystem for text recognition
    text_sys = TextSystem()

    # 步骤编号的正则表达式，匹配大于0的数字后跟一个点
    step_pattern = re.compile(r'(?<!\.\d)\b(?<!\d\.)[1-9]\d*\.(?!\d)')

    # 存储满足步骤编号格式的文本
    step_texts = []

    # 用于存储更新后的图片路径列表
    updated_image_paths = []

    for image_path in image_paths:
        # 使用 OpenCV 读取图像
        img = cv2.imread(image_path)

        # 进行文字识别
        res = text_sys.detect_and_ocr(img)

        # 提取识别到的文本
        matched = False
        for boxed_result in res:
            # 如果识别到的文本满足步骤编号格式，则添加到列表中
            if step_pattern.match(boxed_result.ocr_text):
                step_texts.append(boxed_result.ocr_text)
                matched = True
                break

        # 如果未满足步骤编号格式，删除图片
        if not matched:
            os.remove(image_path)
        else:
            updated_image_paths.append(image_path)  # 如果满足条件，则保留图片路径

    return updated_image_paths


def get_image_text(extracted_images):
    text_sys = TextSystem()
    pattern = r'(\d+)\s*X\s*([A-Z])'
    page_pattern = r'step_(\d+)_'

    letter_counts = defaultdict(int)
    letter_count = defaultdict(list)
    letter_pageNumber = defaultdict(list)

    for image_path in extracted_images:
        img = cv2.imread(image_path)
        res = text_sys.detect_and_ocr(img)

        # 提取页码
        page_match = re.search(page_pattern, image_path)
        if page_match:
            page_number = int(page_match.group(1))

        for boxed_result in res:
            matches = re.findall(pattern, boxed_result.ocr_text)
            for number_str, letter in matches:
                number = int(number_str)
                letter_counts[letter] += number
                letter_count[letter].append(number)

                # 检查页码是否已经记录在列表中
                if page_number not in letter_pageNumber[letter]:
                    letter_pageNumber[letter].append(page_number)

    return dict(letter_counts), dict(letter_count), dict(letter_pageNumber)


# 获取步骤螺丝
def get_step_screw(doc):
    # 提取步骤下方的图像
    extracted_images = extract_images_below_steps(doc)
    extracted_images = recognize_text_in_images(extracted_images)
    letter_counts,letter_count,letter_pageNumber = get_image_text(extracted_images)

    return letter_counts,letter_count,letter_pageNumber


def check_total_and_step(doc):
    count_mismatch = {}  # 数量不匹配的情况
    extra_chars = {}  # 多余的字符
    missing_chars = {}  # 缺少的字符
    result_dict, page_num = get_total_screw(doc)
    letter_counts,letter_count,letter_pageNumber = get_step_screw(doc)

    # 检查两个字典中的数量是否匹配
    for key in letter_counts:
        if key in result_dict:
            if result_dict[key] != letter_counts[key]:
                count_mismatch[key] = {'expected': result_dict[key], 'actual': letter_counts[key]}
                print(f"数量不匹配: {key}, 应有 {result_dict[key]} 个, 实际有 {letter_counts[key]} 个")
        else:
            print(f"多余的字符: {key} 在 result_dict 中不存在")
            extra_chars[key] = letter_counts[key]

    # 检查result_dict是否有letter_counts没有的字符,多余的种类螺丝
    for key in result_dict:
        if key not in letter_counts:
            print(f"缺少的字符: {key} 在 letter_counts 中不存在")
            missing_chars[key] = result_dict[key]

    return count_mismatch, extra_chars, missing_chars, page_num, letter_count, letter_pageNumber, result_dict


# 写回错误信息,返回的是excel
def create_styled_excel(result_dict, count_mismatch, letter_count, letter_pageNumber,file_path):
    # 检查 count_mismatch 是否为空
    if not count_mismatch:
        # 创建 Excel 工作簿和工作表
        wb = Workbook()
        ws = wb.active

        # 创建合并单元格的标题
        ws.merge_cells('A1:E1')
        ws['A1'] = '螺丝对比结果'
        ws.merge_cells('A2:E2')
        ws['A2'] = '正确总数量螺丝'

        # 对标题应用样式
        for cell in ws["1:1"] + ws["2:2"]:
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 添加列标题
        ws.append(['螺丝型号', '螺丝总数量', '步骤螺丝总数量', '步骤螺丝个数', '步骤螺丝页数'])

        # 填充数据
        for letter, total in result_dict.items():
            ws.append([
                letter,
                total,
                total,  # 假设正确数量与总数相同，因为 count_mismatch 为空
                ', '.join(map(str, letter_count.get(letter, []))),
                ', '.join(map(str, letter_pageNumber.get(letter, []))),
            ])

        # 对所有数据单元格应用细边框
        thin = Side(border_style="thin", color="000000")
        for row in ws.iter_rows(min_row=3, max_row=len(result_dict) + 3, min_col=1, max_col=5):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 25

        # 保存文件
        wb.save(file_path)

        return file_path
    else:
        # 将不匹配的键，在result_dict里删除
        for mismatch_key in count_mismatch.keys():
            if mismatch_key in result_dict:
                del result_dict[mismatch_key]  # Remove the key from result_dict

        # 写第一行螺丝对比结果
        wb = Workbook()
        ws = wb.active
        # 记入表格行数
        row_count = 1
        ws.merge_cells(f'A{row_count}:E{row_count}')
        ws['A1'] = '螺丝对比结果'
        # 设置样式和边框
        thin = Side(border_style="thin", color="000000")
        header_font = Font(bold=True, size=12)
        align_center = Alignment(horizontal="center", vertical="center")
        # 应用样式到合并的标题单元格
        for row in ws.iter_rows(min_row=row_count, max_row=row_count, min_col=1, max_col=5):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = align_center
                cell.font = header_font
        row_count += 1

        # 如果 result_dict 不为空，则继续创建 Excel 工作簿和工作表
        if result_dict:
            # 应用样式
            ws.merge_cells(f'A{row_count}:E{row_count}')
            ws[f'A{row_count}'] = '正确总数量螺丝'
            header_font = Font(bold=True, size=12, color="0000FF")
            # 应用样式到合并的标题单元格
            for row in ws.iter_rows(min_row=row_count, max_row=row_count, min_col=1, max_col=5):
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = align_center
                    cell.font = header_font
            row_count += 1
            # 添加列标题
            headers = ['螺丝型号', '螺丝包总数量', '步骤螺丝总数量', '步骤螺丝个数', '步骤螺丝页数']
            ws.append(headers)
            for cell in ws[3]:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = align_center
                cell.font = header_font

            # 添加数据行
            for key, value in result_dict.items():
                row_count += 1
                ws.append([
                    key,
                    value,
                    value,
                    ', '.join(map(str, letter_count.get(key, []))),
                    ', '.join(map(str, letter_pageNumber.get(key, [])))
                ])

            # 应用边框和对齐到数据单元格
            for row in ws.iter_rows(min_row=4, min_col=1, max_col=5, max_row=ws.max_row):
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = align_center

            # 设置列宽
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 25
        # 写错误部分
        # 应用样式
        ws.merge_cells(f'A{row_count}:E{row_count}')
        ws[f'A{row_count}'] = '错误总数量螺丝'
        header_font = Font(bold=True, size=12, color="FF0000")
        # 应用样式到合并的标题单元格
        for row in ws.iter_rows(min_row=row_count, max_row=row_count, min_col=1, max_col=5):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = align_center
                cell.font = header_font
        for mismatch_key, counts in count_mismatch.items():
            expected_count = counts['expected']
            actual_count = counts['actual']
            ws.append([
                mismatch_key,
                expected_count,
                actual_count,
                ', '.join(map(str, letter_count.get(mismatch_key, []))),
                ', '.join(map(str, letter_pageNumber.get(mismatch_key, [])))
            ])
            # 应用边框和对齐到数据单元格
            for row in ws.iter_rows(min_row=row_count, min_col=1, max_col=5, max_row=ws.max_row):
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = align_center
        # 保存文件

        wb.save(file_path)

        return file_path

# 主函数
def check_screw(file):
    doc = fitz.open(stream=BytesIO(file))
    doc.save(PDF_PATH)
    if not os.path.isdir(IMAGE_PATH):
        os.makedirs(IMAGE_PATH)

    count_mismatch, extra_chars, missing_chars, page_num,letter_count,letter_pageNumber,result_dict = check_total_and_step(doc)

    # annotations = {}
    print(f"count_mismatch = {count_mismatch}")
    
    output_excel_path = f'1.xlsx'
    # print(annotations)
    output_excel_path = create_styled_excel(result_dict, count_mismatch, letter_count, letter_pageNumber,output_excel_path)
    # 将文档转换成字节流
    # doc_bytes = doc.write()
    # 将字节流进行base64编码
    # doc_base64 = base64.b64encode(doc_bytes).decode('utf-8')

    doc.close()
    os.remove(CSV_PATH)
    os.remove(PDF_PATH)
    shutil.rmtree(IMAGE_PATH)
    return doc_base64
