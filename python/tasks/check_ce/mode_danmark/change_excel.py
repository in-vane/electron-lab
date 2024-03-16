from openpyxl import load_workbook
from openpyxl.styles import Border, Side

from .get_table_message import all


def change_excel(excel, work_table, message_dict):
    # 加载工作簿并选择工作表
    wb = load_workbook(excel)
    sheet = wb[work_table]

    # 用绿色定义边框样式

    green_border = Border(
        left=Side(style='thick', color='0000FF'),
        right=Side(style='thick', color='0000FF'),
        top=Side(style='thick', color='0000FF'),
        bottom=Side(style='thick', color='0000FF')
    )

    # 如果message_dict中有'CE-sign'键，处理它的值
    if 'CE-sign' in message_dict:
        ce_values = message_dict['CE-sign']  # 获取'CE-sign'对应的值列表
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in ce_values:  # 如果单元格的值在ce_values列表中
                    cell.border = green_border  # 更新边框为绿色

    # 遍历工作表中的每一行，更新其他指定的值的边框
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in message_dict and cell.value != 'CE-sign':  # 排除'CE-sign'键
                red_texts_positions = message_dict[cell.value]  # 获取红色文本位置的列表
                red_texts_count = 0  # 找到的红色文本单元格的计数器

                for row_cell in row:
                    # 检查单元格的字体是否为红色
                    if row_cell.font and row_cell.font.color and row_cell.font.color.type == 'rgb' and row_cell.font.color.value == 'FFFF0000':
                        red_texts_count += 1  # 递增计数器
                        # 如果当前计数在位置列表中，则更新边界
                        if red_texts_count in red_texts_positions:
                            row_cell.border = green_border

    # 保存文件
    wb.save('output.xlsx')


def checkTags(excel, work_table, pdf_path):
    work_table = '例3'
    message_dict = all(excel, work_table, pdf_path)
    change_excel(excel, work_table, message_dict)
