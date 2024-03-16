import openpyxl
import re
from ppocronnx.predict_system import TextSystem
import cv2
import os
import fitz
from .get_similarity import compare_dictionaries


def get_standard_document_as_dict(standard_excel, sheet_name):
    # Load the Excel file
    wb = openpyxl.load_workbook(standard_excel)
    # Access the specified sheet
    sheet = wb[sheet_name]

    red_text_dict = {}  # Create a dictionary to store the results

    # Traverse each row
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        # Assume the table's first column is not empty and starts from the second column
        table_start_column = None
        for cell in row:
            if cell.value is not None and table_start_column is None:
                table_start_column = cell.column

        # If the start of the table is found
        if table_start_column is not None:
            first_cell_value = row[table_start_column - 1].value  # Get the value of the first column of the table
            red_texts = []
            for cell in row[table_start_column:]:
                if cell.font and cell.font.color and cell.font.color.rgb == 'FFFF0000':  # Look for red font
                    red_texts.append(cell.value)  # Add the red font value to the list

            if red_texts:
                # If the key is already in the dictionary, append the new red text value
                if first_cell_value in red_text_dict:
                    red_text_dict[first_cell_value].extend(red_texts)
                else:
                    # Otherwise, create a new key in the dictionary
                    red_text_dict[first_cell_value] = red_texts

    # Close the workbook
    wb.close()
    print(red_text_dict)
    red_text_dict = update_key_standard_dict(red_text_dict)
    return red_text_dict


def update_key_standard_dict(data_dict):
    """
        将字典中值为xxxx-xx的键设为CE-sign
    :param data_dict:一个字典
    :return:修改后的字典
    """
    # Define the regex pattern for 'xxxx-xx'
    ce_sign_pattern = re.compile(r'\b\d{4}-\d{2}\b')

    # Create a new dictionary to store updated results
    updated_dict = {}

    # Iterate over items in the original dictionary
    for key, values in data_dict.items():
        # Assume each key only has one value in the list
        if values and ce_sign_pattern.match(values[0]):
            # If the value matches the pattern, change the key to 'CE-sign'
            updated_dict['CE-sign'] = values
        else:
            # Otherwise, keep the original key-value pair
            updated_dict[key] = values

    return updated_dict


def pdf_to_images(pdf_path):
    # 确保输出文件夹存在
    output_folder = "images"
    os.makedirs(output_folder, exist_ok=True)

    # 打开PDF文件
    doc = fitz.open(pdf_path)

    image_paths = []  # 用于存储生成的图片路径

    # 遍历PDF中的每一页
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)  # 加载当前页
        pix = page.get_pixmap()  # 将当前页渲染成一个位图
        output_path = f"{output_folder}/page_{page_num + 1}.png"  # 定义输出图片的路径
        pix.save(output_path)  # 保存图片
        image_paths.append(output_path)  # 将路径添加到列表中

    doc.close()  # 关闭PDF文档

    return image_paths  # 返回图片路径列表


def extract_text_from_image(image_path):
    text_sys = TextSystem()
    img = cv2.imread(image_path)
    res = text_sys.detect_and_ocr(img)

    extracted_text = ""
    for boxed_result in res:
        extracted_text += "{}\n".format(boxed_result.ocr_text)
    print(extracted_text)
    extracted_text = extract_key_value_pairs(extracted_text)
    return extracted_text


def extract_key_value_pairs(text):
    lines = text.split('\n')
    key_value_pairs = {}
    previous_key = None

    for i, line in enumerate(lines):
        if ':' in line:
            parts = line.split(':')
            key = parts[0].strip()
            value = ':'.join(parts[1:]).strip()
            if value:
                # If there is a value following the colon, use it directly
                key_value_pairs[key] = value
                previous_key = None  # Reset previous key as the current line is self-contained
            else:
                # If there isn't a value following the colon, prepare to use the next line
                previous_key = key
                if i + 1 < len(lines) and ':' not in lines[i + 1]:
                    value = lines[i + 1].strip()
                    key_value_pairs[key] = value
                    i += 1  # Increment the index to skip the next line as it's used here
        elif previous_key and ':' not in line:
            # Assign the line as value to the last found key if the line has no colon
            key_value_pairs[previous_key] = line.strip()
            previous_key = None  # Reset previous key as its value has been assigned
    key_value_pairs = {key: [value] for key, value in key_value_pairs.items()}
    return key_value_pairs


def all(excel_path, work_table, pdf_path):
    # 假设Excel文件已经保存在以下路径

    # 调用函数并打印结果
    red_text_data = get_standard_document_as_dict(excel_path, work_table)
    print(red_text_data)

    # 调用函数并传入PDF文件的路径
    # 指定图像文件路径
    list = pdf_to_images(pdf_path)
    for image_path in list:
        table_data = extract_text_from_image(image_path)
    print(table_data)
    red_text_data = compare_dictionaries(red_text_data, table_data)
    print(red_text_data)
    return red_text_data

# all()
