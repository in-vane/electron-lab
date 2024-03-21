import os
import base64
from io import BytesIO

import fitz
import openpyxl
from openpyxl.styles import Border, Side

import os
import jpype


from .get_table_message import all

CURRENT_PATH = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(CURRENT_PATH, "temp.xlsx")
IMAGE_PATH = os.path.join(CURRENT_PATH, 'temp.png')
PDF_PATH1 = os.path.join(CURRENT_PATH, '2.pdf')
PDF_PATH2 = os.path.join(CURRENT_PATH, 'temp.pdf')


def change_excel(wb, work_table, message_dict):
    """
    该函数主要是为了，呈现错误信息到吉盛标准ce表上
    :param excel: 吉森标准ce表
    :param work_table: ce表中的工作表
    :param message_dict: 错误信息
    :return:
    """
    # 加载工作簿并选择工作表
    # wb = load_workbook(excel)
    sheet = wb[work_table]

    # 用绿色定义边框样式
    # 用绿色定义边框样式
    green_border = Border(
        left=Side(style='thick', color='0000FF'),
        right=Side(style='thick', color='0000FF'),
        top=Side(style='thick', color='0000FF'),
        bottom=Side(style='thick', color='0000FF')
    )

    # 如果message_dict中有'CE-sign'键，处理它的值
    if 'CE-sign' in message_dict:
        ce_values = message_dict['CE-sign']  # 获取'CE-sign'对应的值列表
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in ce_values:  # 如果单元格的值在ce_values列表中
                    cell.border = green_border  # 更新边框为绿色

    # 遍历工作表中的每一行，更新其他指定的值的边框
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in message_dict and cell.value != 'CE-sign':  # 排除'CE-sign'键
                red_texts_positions = message_dict[cell.value]  # 获取红色文本位置的列表
                red_texts_count = 0  # 找到的红色文本单元格的计数器

                for row_cell in row:
                    # 检查单元格的字体是否为红色
                    if row_cell.font and row_cell.font.color and row_cell.font.color.type == 'rgb' and row_cell.font.color.value == 'FFFF0000':
                        red_texts_count += 1  # 递增计数器
                        # 如果当前计数在位置列表中，则更新边界
                        if red_texts_count in red_texts_positions:
                            row_cell.border = green_border

    # 保存文件
    # wb.save(EXCEL_PATH)


# 将ecxel转化pdf
def convert_excel_sheet_to_pdf(excel_file, sheet_name):
    # 启动JVM
    jpype.startJVM()
    from asposecells.api import Workbook, PdfSaveOptions

    # 加载Excel文档
    workbook = Workbook(excel_file)

    # 获取所有工作表的集合
    worksheets = workbook.getWorksheets()

    # 遍历所有工作表
    for sheet in worksheets:
        print(f"excel中的工作表{sheet}")
        # 如果工作表不是要转换的工作表，将其隐藏
        if sheet.getName() != sheet_name:
            sheet.setVisible(False)

    # 设置PDF保存选项
    saveOptions = PdfSaveOptions()
    saveOptions.setOnePagePerSheet(True)

    # 保存为PDF
    workbook.save(PDF_PATH2, saveOptions)

    # 关闭JVM
    jpype.shutdownJVM()

# 将pdf的第一页转化图片
def convert_pdf_page_to_image(pdf_path, image_path, page_number=0):
    """
    Convert the specified page of a PDF file into an image.

    :param pdf_path: Path to the PDF file.
    :param image_path: Path where the image will be saved.
    :param page_number: The number of the page to convert (0-based).
    """
    # 打开PDF文件
    doc = fitz.open(pdf_path)

    # 选择PDF的第一页
    page = doc.load_page(page_number)

    # 将选中的页面转换为图片（pix）
    pix = page.get_pixmap()

    # 将图片保存为文件
    pix.save(image_path)

    # 关闭文档
    doc.close()


def checkTags(excel_file, pdf_file):
    work_table = '例2'
    # doc = fitz.open(pdf_file)
    # wb = openpyxl.load_workbook(excel_file)
    doc = fitz.open(stream=BytesIO(pdf_file))
    doc.save(PDF_PATH1)
    wb = openpyxl.load_workbook(filename=BytesIO(excel_file))
    wb.save(EXCEL_PATH)
    message_dict = all(wb, work_table, doc, PDF_PATH1)
    change_excel(wb, work_table, message_dict)
    # 将excel转化为pdf，保存到PDF_PATH2
    convert_excel_sheet_to_pdf(EXCEL_PATH, work_table)
    # 将excel转化为pdf，保存到PDF_PATG
    convert_pdf_page_to_image(PDF_PATH2, IMAGE_PATH)

    with open(IMAGE_PATH, "rb") as image_file:
        image_data = image_file.read()
    # 将二进制流编码为Base64字符串
    base64_encoded_data = base64.b64encode(image_data)

    # 将Base64字节对象转换为字符串
    image_base64 = base64_encoded_data.decode('utf-8')
    print(image_base64)
    doc.close()
    os.remove(EXCEL_PATH)
    os.remove(IMAGE_PATH)
    os.remove(PDF_PATH1)
    os.remove(PDF_PATH2)
    # Close the workbook
    #     # wb.close()

    return image_base64


# checkTags('2.xlsx','2.pdf')